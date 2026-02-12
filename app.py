import streamlit as st
import openpyxl
import io
from collections import defaultdict
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries

st.set_page_config(page_title="Excel Обработчик", page_icon="📊")

st.title("📊 Обработка Excel файлов")
st.markdown("---")

# ==================== РЕЖИМ 1: БЕЛЬГИЯ (29 символов, депеши) ====================
def add_statistics_table_belgium(ws):
    """Добавляет таблицу статистики для режима БЕЛЬГИЯ"""
    try:
        # Находим строку с "общее количество"
        summary_row = None
        for row in range(1, ws.max_row + 1):
            if ws[f'A{row}'].value == "общее количество":
                summary_row = row
                break
        
        if not summary_row:
            return
        
        # Создаем таблицу через 1 строку
        table_start_row = summary_row + 2
        
        # Заголовки таблицы
        headers = [
            "номер депеши",
            "кол-во всего",
            "кол-во посылки", 
            "кол-во мешки",
            "вес всего",
            "вес посылки",
            "вес мешки"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws[f'{col_letter}{table_start_row}'] = header
        
        # Собираем данные
        depesh_codes = set()
        data_by_depesh = defaultdict(lambda: {'pos_count': 0, 'mesh_count': 0, 'pos_weight': 0.0, 'mesh_weight': 0.0})
        
        # Анализируем данные Посылок (A-C)
        for row in range(2, ws.max_row + 1):
            code = ws[f'A{row}'].value
            depesh = ws[f'B{row}'].value
            weight = ws[f'C{row}'].value
            
            if code and depesh and weight is not None:
                if isinstance(depesh, str) and len(depesh) == 4 and depesh.isdigit():
                    depesh_codes.add(depesh)
                    data_by_depesh[depesh]['pos_count'] += 1
                    data_by_depesh[depesh]['pos_weight'] += float(weight)
        
        # Анализируем данные Мешков (D-F)
        for row in range(2, ws.max_row + 1):
            code = ws[f'D{row}'].value
            depesh = ws[f'E{row}'].value
            weight = ws[f'F{row}'].value
            
            if code and depesh and weight is not None:
                if isinstance(depesh, str) and len(depesh) == 4 and depesh.isdigit():
                    depesh_codes.add(depesh)
                    data_by_depesh[depesh]['mesh_count'] += 1
                    data_by_depesh[depesh]['mesh_weight'] += float(weight)
        
        # Заполняем таблицу
        sorted_depesh_codes = sorted(depesh_codes)
        for idx, depesh in enumerate(sorted_depesh_codes, start=1):
            table_row = table_start_row + idx
            data = data_by_depesh[depesh]
            
            ws[f'A{table_row}'] = depesh
            ws[f'B{table_row}'] = data['pos_count'] + data['mesh_count']
            ws[f'C{table_row}'] = data['pos_count']
            ws[f'D{table_row}'] = data['mesh_count']
            ws[f'E{table_row}'] = round(data['pos_weight'] + data['mesh_weight'], 1)
            ws[f'F{table_row}'] = round(data['pos_weight'], 1)
            ws[f'G{table_row}'] = round(data['mesh_weight'], 1)
            
    except Exception as e:
        st.error(f"Ошибка при добавлении статистики: {e}")

def process_belgium(file_bytes):
    """Обработка для БЕЛЬГИИ (29 символов, депеши)"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        
        # ШАГ 1: Вставить два столбца между A и B
        ws.insert_cols(2, 2)
        
        # ШАГ 2: Заполнить столбец B (номер депеши из A)
        ws['B1'] = "номер депеши"
        for row in range(2, ws.max_row + 1):
            cell_a = ws[f'A{row}'].value
            if cell_a and len(str(cell_a)) == 29:
                ws[f'B{row}'] = str(cell_a)[16:20]
        
        # ШАГ 3: Заполнить столбец C (вес из A)
        ws['C1'] = "вес"
        for row in range(2, ws.max_row + 1):
            cell_a = ws[f'A{row}'].value
            if cell_a and len(str(cell_a)) == 29:
                try:
                    ws[f'C{row}'] = int(str(cell_a)[-4:]) / 10
                except:
                    pass
        
        # ШАГ 4: Заполнить столбец E (номер депеши из D)
        ws['E1'] = "номер депеши"
        for row in range(2, ws.max_row + 1):
            cell_d = ws[f'D{row}'].value
            if cell_d and len(str(cell_d)) == 29:
                ws[f'E{row}'] = str(cell_d)[16:20]
        
        # ШАГ 5: Заполнить столбец F (вес из D)
        ws['F1'] = "вес"
        for row in range(2, ws.max_row + 1):
            cell_d = ws[f'D{row}'].value
            if cell_d and len(str(cell_d)) == 29:
                try:
                    ws[f'F{row}'] = int(str(cell_d)[-4:]) / 10
                except:
                    pass
        
        # Найти последнюю строку с кодом
        last_row = 1
        for row in range(2, ws.max_row + 1):
            cell_a = ws[f'A{row}'].value
            cell_d = ws[f'D{row}'].value
            if (cell_a and len(str(cell_a)) == 29) or (cell_d and len(str(cell_d)) == 29):
                last_row = row
        
        # ШАГ 6: Общий вес
        start_row = last_row + 4
        ws[f'A{start_row}'] = "общий вес"
        total_weight = 0
        for row in range(2, ws.max_row + 1):
            weight_c = ws[f'C{row}'].value
            weight_f = ws[f'F{row}'].value
            if isinstance(weight_c, (int, float)):
                total_weight += weight_c
            if isinstance(weight_f, (int, float)):
                total_weight += weight_f
        ws[f'B{start_row}'] = round(total_weight, 1)
        
        # ШАГ 7: Общее количество
        ws[f'A{start_row + 1}'] = "общее количество"
        total_count = 0
        for row in range(2, ws.max_row + 1):
            cell_a = ws[f'A{row}'].value
            cell_d = ws[f'D{row}'].value
            if cell_a and len(str(cell_a)) == 29:
                total_count += 1
            if cell_d and len(str(cell_d)) == 29:
                total_count += 1
        ws[f'B{start_row + 1}'] = total_count
        
        # Добавляем таблицу статистики
        add_statistics_table_belgium(ws)
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"Ошибка при обработке: {e}")
        return None

# ==================== РЕЖИМ 2: ГОЛЛАНДИЯ (Pallet, блоки, итоги) ====================
def unmerge_all_cells(ws):
    """Разъединяет все объединённые ячейки"""
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

def process_holland_blocks(ws):
    """Обработка блоков строк для ГОЛЛАНДИИ"""
    max_row = ws.max_row
    current_row = 1
    block_start = None
    blocks = []
    bold_font = Font(bold=True)
    
    # Идентифицируем блоки строк по слову "Pallet"
    while current_row <= max_row:
        cell_a = ws.cell(row=current_row, column=1)
        cell_value = cell_a.value
        
        if cell_value and isinstance(cell_value, str) and 'Pallet' in cell_value:
            if block_start is None:
                block_start = current_row
            elif block_start is not None and current_row > block_start:
                blocks.append((block_start, current_row - 1))
                block_start = current_row
        
        current_row += 1
    
    if block_start is not None:
        blocks.append((block_start, max_row))
    
    # Если не нашли блоки через Pallet, ищем по кодам
    if not blocks:
        current_row = 1
        block_start = None
        
        while current_row <= max_row:
            cell_a = ws.cell(row=current_row, column=1)
            cell_c = ws.cell(row=current_row, column=3)
            
            has_code = False
            if cell_a.value and isinstance(cell_a.value, str) and len(str(cell_a.value).strip()) >= 12:
                has_code = True
            elif cell_c.value and isinstance(cell_c.value, str) and len(str(cell_c.value).strip()) >= 12:
                has_code = True
            
            if has_code and block_start is None:
                block_start = current_row
            elif block_start is not None and not has_code and current_row > block_start:
                blocks.append((block_start, current_row - 1))
                block_start = None
            
            current_row += 1
        
        if block_start is not None:
            blocks.append((block_start, max_row))
    
    # Смещение для вставки строк
    row_offset = 0
    
    # Обрабатываем каждый блок
    for block_num, (start_row, end_row) in enumerate(blocks):
        start_row += row_offset
        end_row += row_offset
        
        # Проверяем, содержит ли блок коды нужной длины (13-29 символов)
        has_valid_codes = False
        
        for row in range(start_row, end_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_c = ws.cell(row=row, column=3)
            
            if cell_a.value and isinstance(cell_a.value, str):
                code = str(cell_a.value).strip()
                if 13 <= len(code) <= 29:
                    has_valid_codes = True
            
            if cell_c.value and isinstance(cell_c.value, str):
                code = str(cell_c.value).strip()
                if 13 <= len(code) <= 29:
                    has_valid_codes = True
        
        if not has_valid_codes:
            continue
        
        # Добавляем строку "КОЛ-ВО ОТПРАВЛЕНИЙ" после блока
        ws.insert_rows(end_row + 1)
        row_offset += 1
        count_row = end_row + 1
        
        ws.cell(row=count_row, column=1).value = "КОЛ-ВО ОТПРАВЛЕНИЙ"
        ws.cell(row=count_row, column=1).font = bold_font
        
        # Подсчет количества кодов в блоке (12-29 символов)
        code_count = 0
        for row in range(start_row, end_row + 1):
            for col in [1, 3]:
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    code = str(cell.value).strip()
                    if 12 <= len(code) <= 29:
                        code_count += 1
        
        ws.cell(row=count_row, column=2).value = code_count
        
        # Добавляем строку "ВЕС ОТПРАВЛЕНИЙ"
        ws.insert_rows(count_row + 1)
        row_offset += 1
        weight_row = count_row + 1
        
        ws.cell(row=weight_row, column=1).value = "ВЕС ОТПРАВЛЕНИЙ"
        ws.cell(row=weight_row, column=1).font = bold_font
        
        # Подсчет суммы весов в блоке
        weight_sum = 0
        for row in range(start_row, end_row + 1):
            for col in [2, 4]:
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    weight_sum += cell.value
        
        ws.cell(row=weight_row, column=2).value = round(weight_sum, 1)
        
        # Обновляем end_row для следующей итерации
        end_row += 2
    
    # Удаляем строки с "КОЛ-ВО ОТПРАВЛЕНИЙ" и "ВЕС ОТПРАВЛЕНИЙ", которые идут сразу после Pallet
    rows_to_delete = []
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        if cell_a.value in ["КОЛ-ВО ОТПРАВЛЕНИЙ", "ВЕС ОТПРАВЛЕНИЙ"]:
            if row > 1:
                prev_cell_a = ws.cell(row=row-1, column=1)
                if prev_cell_a.value and isinstance(prev_cell_a.value, str) and 'Pallet' in prev_cell_a.value:
                    rows_to_delete.append(row)
    
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
    
    # Добавляем итоговые строки в строках 3 и 4
    bold_font = Font(bold=True)
    
    for row in [3, 4]:
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None
    
    # "ВСЕГО КОЛ-ВО ОТПРАВЛЕНИЙ" в строке 3, столбец H
    ws.cell(row=3, column=8).value = "ВСЕГО КОЛ-ВО ОТПРАВЛЕНИЙ"
    ws.cell(row=3, column=8).font = bold_font
    
    total_count = 0
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)
        if cell_a.value == "КОЛ-ВО ОТПРАВЛЕНИЙ" and isinstance(cell_b.value, (int, float)):
            total_count += cell_b.value
    
    ws.cell(row=3, column=9).value = total_count
    
    # "ОБЩИЙ ВЕС ОТПРАВЛЕНИЙ" в строке 4, столбец H
    ws.cell(row=4, column=8).value = "ОБЩИЙ ВЕС ОТПРАВЛЕНИЙ"
    ws.cell(row=4, column=8).font = bold_font
    
    total_weight = 0
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)
        if cell_a.value == "ВЕС ОТПРАВЛЕНИЙ" and isinstance(cell_b.value, (int, float)):
            total_weight += cell_b.value
    
    ws.cell(row=4, column=9).value = round(total_weight, 1)

def process_holland(file_bytes):
    """Обработка для ГОЛЛАНДИИ (Pallet, блоки, итоги)"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        
        # Разъединяем все объединенные ячейки
        unmerge_all_cells(ws)
        
        # Шаг 1: Заменяем "Pal" на "Pallet" в первом столбце
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str):
                cell_value = str(cell.value)
                if cell_value.startswith('Pal '):
                    cell.value = cell_value.replace('Pal ', 'Pallet ')
                elif cell_value == 'Pal':
                    cell.value = 'Pallet'
        
        # Сохраняем текущие данные столбца B (теперь он станет столбцом C)
        max_row = ws.max_row
        old_column_b = []
        
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=2)
            old_column_b.append(cell.value)
            cell.value = None
        
        # Вставляем новый столбец между A и B
        ws.insert_cols(2)
        
        # Восстанавливаем старые данные столбца B в новый столбец C
        for row, value in enumerate(old_column_b, 1):
            if value is not None:
                ws.cell(row=row, column=3).value = value
        
        # Обрабатываем коды в столбце A -> результат в столбец B
        for row in range(1, max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            
            if cell_a.value and isinstance(cell_a.value, str):
                code = str(cell_a.value).strip()
                if len(code) >= 28:
                    try:
                        last_4_digits = code[-4:]
                        if last_4_digits.isdigit():
                            ws.cell(row=row, column=2).value = int(last_4_digits) / 10
                    except:
                        pass
        
        # Обрабатываем коды в столбце C -> результат в столбец D
        if ws.max_column < 4:
            ws.insert_cols(4)
        
        for row in range(1, max_row + 1):
            cell_c = ws.cell(row=row, column=3)
            
            if cell_c.value and isinstance(cell_c.value, str):
                code = str(cell_c.value).strip()
                if len(code) >= 28:
                    try:
                        last_4_digits = code[-4:]
                        if last_4_digits.isdigit():
                            ws.cell(row=row, column=4).value = int(last_4_digits) / 10
                    except:
                        pass
        
        # Обрабатываем блоки строк
        process_holland_blocks(ws)
        
        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"Ошибка при обработке ГОЛЛАНДИЯ: {e}")
        return None

# ==================== ИНТЕРФЕЙС ПОЛЬЗОВАТЕЛЯ ====================

# Выбор режима обработки
mode = st.radio(
    "Выберите тип обработки:",
    ["🇧🇪 БЕЛЬГИЯ (29 символов, депеши, статистика)", 
     "🇳🇱 ГОЛЛАНДИЯ (Pallet, блоки, итоги)"],
    horizontal=True
)

st.markdown("---")

# Загрузка файла
uploaded_file = st.file_uploader("Выберите Excel файл", type=['xlsx'])

if uploaded_file is not None:
    st.info(f"📁 Загружен: {uploaded_file.name} ({uploaded_file.size / 1024:.1f} KB)")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🚀 ОБРАБОТАТЬ ФАЙЛ", type="primary", use_container_width=True):
            with st.spinner("⏳ Идет обработка..."):
                file_bytes = uploaded_file.getvalue()
                
                if "БЕЛЬГИЯ" in mode:
                    processed_file = process_belgium(file_bytes)
                    file_prefix = "BELGIUM"
                else:
                    processed_file = process_holland(file_bytes)
                    file_prefix = "HOLLAND"
                
                if processed_file:
                    st.success("✅ Файл успешно обработан!")
                    
                    with col2:
                        st.download_button(
                            label="📥 СКАЧАТЬ ОБРАБОТАННЫЙ ФАЙЛ",
                            data=processed_file,
                            file_name=f"{file_prefix}_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    st.balloons()
else:
    st.info("👆 Загрузите файл .xlsx для начала обработки")

st.markdown("---")
st.markdown("""
### 📌 Описание режимов:

**🇧🇪 БЕЛЬГИЯ:**
- Для файлов с кодами 29 символов
- Добавляет номера депеш (16-20 символы)
- Добавляет вес (последние 4 цифры / 10)
- Считает общий вес и общее количество
- Создает таблицу статистики по номерам депеш

**🇳🇱 ГОЛЛАНДИЯ:**
- Для файлов с Pallet и кодами 13-29 символов
- Заменяет "Pal" на "Pallet"
- Добавляет вес из последних 4 цифр
- Группирует в блоки
- Добавляет КОЛ-ВО ОТПРАВЛЕНИЙ и ВЕС ОТПРАВЛЕНИЙ
- Считает ВСЕГО КОЛ-ВО и ОБЩИЙ ВЕС
""")
